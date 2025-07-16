import os
import io
from datetime import datetime
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from flask import Flask, render_template, request, redirect, url_for, flash, Response
from sqlalchemy import or_, and_
from models import db, Ingrediente, FichaTecnica, FichaTecnicaIngrediente, Forma, Configuracao, Compra, Venda, User
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import User
from xhtml2pdf import pisa
from flask import send_file
from io import BytesIO
import unicodedata
from flask_migrate import Migrate




app = Flask(__name__)


app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'uma-chave-secreta-muito-forte-e-aleatoria')

# Configuração do Banco de Dados
database_uri = os.environ.get('DATABASE_URL')
if database_uri and database_uri.startswith("postgres://"):
    database_uri = database_uri.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_uri or 'sqlite:///confeitando_com_artes.db'
migrate = Migrate(app, db) 
db.init_app(app)

# --- Configuração do Login ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Por favor, faça o login para acessar esta página."
login_manager.login_message_category = "info"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Comandos do Terminal ---
@app.cli.command('init-db')
def init_db_command():
    db.create_all()
    print('Banco de dados inicializado com sucesso!')

@app.cli.command("create-user")
def create_user():
    """Cria o usuário administrador inicial."""
    username = input("Digite o nome de usuário: ")
    password = input("Digite a senha: ")
    user = User.query.filter_by(username=username).first()
    if user:
        print("Usuário já existe.")
        return
    new_user = User(username=username)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    print(f"Usuário {username} criado com sucesso!")

#LOGIN
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # LINHA DE DEPURAÇÃO: VAMOS IMPRIMIR O QUE VEM DO FORMULÁRIO
        print("Dados do formulário recebidos:", request.form)
        
        user = User.query.filter_by(username=request.form.get('username')).first()

        if user and user.check_password(request.form.get('password')):
            remember_me = request.form.get('remember')
            
            # OUTRA LINHA DE DEPURAÇÃO
            print(f"Checkbox 'remember' foi marcada? {'Sim' if remember_me else 'Não'}")

            login_user(user, remember=(remember_me is not None))
            
            flash('Login realizado com sucesso!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuário ou senha inválidos.', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- ROTAS PRINCIPAIS E DE GESTÃO ---
# Em app.py

@app.route('/')
@login_required
def index():
    # --- Lógica para o Gráfico de Estoque (Mantida) ---
    ingredientes_estoque_grafico = Ingrediente.query.order_by(Ingrediente.quantidade_estoque.desc()).limit(5).all()
    labels_grafico = [ing.nome for ing in ingredientes_estoque_grafico]
    dados_grafico = [ing.quantidade_estoque for ing in ingredientes_estoque_grafico]
    
    # --- Contagem de Fichas (Mantida) ---
    total_fichas = FichaTecnica.query.count()

    # --- LÓGICA DE PAGINAÇÃO PARA OS ALERTAS (MODIFICADA) ---
    
    # 1. Pega o número da página para os alertas a partir da URL
    page_alertas = request.args.get('page_alertas', 1, type=int)

    # 2. Define as mesmas condições de antes para estoque baixo
    condicao_critica = db.or_(
        db.and_(Ingrediente.unidade_medida.in_(['g', 'ml']), Ingrediente.quantidade_estoque <= 100),
        db.and_(Ingrediente.unidade_medida == 'unid', Ingrediente.quantidade_estoque <= 5)
    )
    condicao_alerta = db.or_(
        db.and_(Ingrediente.unidade_medida.in_(['g', 'ml']), Ingrediente.quantidade_estoque > 100, Ingrediente.quantidade_estoque <= 500),
        db.and_(Ingrediente.unidade_medida == 'unid', Ingrediente.quantidade_estoque > 5, Ingrediente.quantidade_estoque <= 15)
    )
    
    # 3. Cria a query base para os alertas
    query_alertas = Ingrediente.query.filter(db.or_(condicao_critica, condicao_alerta)).order_by(Ingrediente.quantidade_estoque.asc())
    
    # 4. Pagina os resultados, mostrando 5 por página
    paginacao_alertas = query_alertas.paginate(page=page_alertas, per_page=5, error_out=False)

    # 5. Passa o objeto de paginação para o template, em vez da lista completa
    return render_template(
        'index.html', 
        total_fichas=total_fichas, 
        labels_grafico=labels_grafico, 
        dados_grafico=dados_grafico, 
        paginacao_alertas=paginacao_alertas  # Nome da variável alterado para refletir a paginação
    )
#para salvar sem acento no banco de dados assim nao aceita duplicação com o mesmo nome com ou sem acento nos ingredientes.#


def normalize_text(text):
    """Remove acentos e converte para minúsculas."""
    if not text:
        return ""
    # Normaliza para o formato NFD (Canonical Decomposition) e remove os caracteres diacríticos (acentos)
    nfkd_form = unicodedata.normalize('NFD', text)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).lower()











@app.route('/ingredientes', methods=['GET', 'POST'])
@login_required
def gerenciar_ingredientes():
    if request.method == 'POST':
        nome_original = request.form.get('nome', '').strip()
        
        # 1. NORMALIZA O NOME PARA VALIDAÇÃO
        nome_normalizado = normalize_text(nome_original)

        # 2. VALIDAÇÃO MELHORADA (IGNORA ACENTOS E MAIÚSCULAS/MINÚSCULAS)
        if Ingrediente.query.filter(Ingrediente.nome_normalizado == nome_normalizado).first():
            flash(f'Erro: O ingrediente "{nome_original}" já está cadastrado.', 'danger')
            return redirect(url_for('gerenciar_ingredientes', focus='nome'))

        # --- Lógica de negócio (mantida como a sua original, mas mais segura com .get()) ---
        tipo_preco = request.form.get('tipo_preco')
        preco_informado = float(request.form.get('preco', 0) or 0)
        unidades_compradas = int(request.form.get('unidades_compradas', 1) or 1)
        tamanho_unidade = float(request.form.get('tamanho_unidade', 0) or 0)
        
        if tipo_preco == 'total':
            preco_total_pago = preco_informado
        else: # tipo_preco == 'unitario'
            preco_total_pago = preco_informado * unidades_compradas
        
        estoque_inicial = tamanho_unidade * unidades_compradas
        custo_inicial = preco_total_pago / estoque_inicial if estoque_inicial > 0 else 0
        
        # 3. SALVA O NOME ORIGINAL E O NOME NORMALIZADO
        novo_ingrediente = Ingrediente(
            nome=nome_original, 
            nome_normalizado=nome_normalizado, # Campo novo sendo salvo
            marca=request.form.get('marca'), 
            unidade_medida=request.form.get('unidade_medida'), 
            quantidade_estoque=estoque_inicial, 
            custo_medio_por_unidade_base=custo_inicial
        )

        primeira_compra = Compra(
            preco_total_lote=preco_total_pago, 
            unidades_compradas=unidades_compradas, 
            tamanho_unidade=tamanho_unidade, 
            ingrediente=novo_ingrediente
        )
        
        db.session.add(novo_ingrediente)
        db.session.add(primeira_compra)
        db.session.commit()
        
        flash(f'Ingrediente "{nome_original}" cadastrado com sucesso!', 'success')
        
        # 4. REDIRECIONA COM PARÂMETRO DE FOCO
        return redirect(url_for('gerenciar_ingredientes', focus='nome'))

    # --- Lógica para GET (exibição da página) ---
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '', type=str)
    
    query = Ingrediente.query
    
    if q:
        # 5. BUSCA MELHORADA (IGNORA ACENTOS E MAIÚSCULAS/MINÚSCULAS)
        q_normalizado = f"%{normalize_text(q)}%"
        query = query.filter(Ingrediente.nome_normalizado.ilike(q_normalizado))
        
    paginacao = query.order_by(Ingrediente.nome).paginate(page=page, per_page=5, error_out=False)
    ingredientes_da_pagina = paginacao.items
    
    return render_template('ingredientes.html', 
                           ingredientes=ingredientes_da_pagina, 
                           paginacao=paginacao, 
                           q=q)

@app.route('/ingrediente/adicionar-estoque/<int:ingrediente_id>', methods=['POST'])
@login_required
def adicionar_estoque(ingrediente_id):
    ingrediente = Ingrediente.query.get_or_404(ingrediente_id)
    preco_lote = float(request.form.get('preco_lote', 0) or 0)
    unidades_lote = int(request.form.get('unidades_lote', 1) or 1)
    tamanho_unidade_lote = float(request.form.get('tamanho_unidade_lote', 0) or 0)
    if preco_lote <= 0 or tamanho_unidade_lote <= 0:
        flash('Valores inválidos para a nova compra.', 'danger')
        return redirect(url_for('gerenciar_ingredientes'))
    nova_compra = Compra(preco_total_lote=preco_lote, unidades_compradas=unidades_lote, tamanho_unidade=tamanho_unidade_lote, ingrediente_id=ingrediente_id)
    db.session.add(nova_compra)
    estoque_antigo_total = ingrediente.quantidade_estoque
    valor_total_estoque_antigo = estoque_antigo_total * ingrediente.custo_medio_por_unidade_base
    quantidade_adicionada = tamanho_unidade_lote * unidades_lote
    novo_estoque_total = estoque_antigo_total + quantidade_adicionada
    novo_valor_total = valor_total_estoque_antigo + preco_lote
    novo_custo_medio = novo_valor_total / novo_estoque_total if novo_estoque_total > 0 else 0
    ingrediente.quantidade_estoque = novo_estoque_total
    ingrediente.custo_medio_por_unidade_base = novo_custo_medio
    db.session.commit()
    flash(f'Estoque de "{ingrediente.nome}" atualizado! Novo custo médio: R$ {novo_custo_medio:.4f}', 'success')
    return redirect(url_for('gerenciar_ingredientes'))

@app.route('/ingrediente/edit/<int:ingrediente_id>', methods=['GET', 'POST'])
@login_required
def editar_ingrediente(ingrediente_id):
    ingrediente = Ingrediente.query.get_or_404(ingrediente_id)
    if request.method == 'POST':
        novo_nome = request.form['nome'].strip()
        outro_ingrediente = Ingrediente.query.filter(Ingrediente.id != ingrediente_id, db.func.lower(Ingrediente.nome) == db.func.lower(novo_nome)).first()
        if outro_ingrediente:
            flash('Erro: Já existe outro ingrediente com este nome.', 'danger')
        else:
            ingrediente.nome = novo_nome
            ingrediente.marca = request.form['marca']
            db.session.commit()
            flash('Ingrediente atualizado com sucesso!', 'success')
            return redirect(url_for('gerenciar_ingredientes'))
    return render_template('editar_ingrediente.html', ingrediente=ingrediente)

@app.route('/ingrediente/delete/<int:ingrediente_id>', methods=['POST'])
@login_required
def deletar_ingrediente(ingrediente_id):
    ingrediente = Ingrediente.query.get_or_404(ingrediente_id)
    db.session.delete(ingrediente)
    db.session.commit()
    flash(f'Ingrediente "{ingrediente.nome}" foi excluído.', 'success')
    return redirect(url_for('gerenciar_ingredientes'))

@app.route('/ingrediente/<int:ingrediente_id>/historico')
@login_required
def historico_compras(ingrediente_id):
    ingrediente = Ingrediente.query.get_or_404(ingrediente_id)
    historico = Compra.query.filter_by(ingrediente_id=ingrediente.id).order_by(Compra.data_compra.desc()).all()
    return render_template('historico_compras.html', ingrediente=ingrediente, historico=historico)

#para fazer compras dos ingredientes baixos
@app.route('/lista-de-compras')
@login_required
def lista_de_compras():
    # Define os limites de estoque baixo
    limite_unidade = 5
    limite_peso_volume = 100 # para g e ml

    # --- CONSULTA CORRIGIDA ---
    # Busca ingredientes que atendem a uma das duas condições abaixo
    ingredientes_a_comprar = Ingrediente.query.filter(
        db.or_(
            # Condição 1: Para itens contados por unidade
            db.and_(
                Ingrediente.unidade_medida == 'unid', 
                Ingrediente.quantidade_estoque <= limite_unidade
            ),
            # Condição 2: Para itens por peso ou volume
            db.and_(
                Ingrediente.unidade_medida.in_(['g', 'ml']), 
                Ingrediente.quantidade_estoque <= limite_peso_volume
            )
        )
    ).order_by(Ingrediente.nome).all()

    return render_template('lista_de_compras.html', ingredientes=ingredientes_a_comprar)




@app.route('/formas', methods=['GET', 'POST'])
@login_required
def gerenciar_formas():
    if request.method == 'POST':
        descricao = request.form['descricao'].strip()
        if Forma.query.filter(db.func.lower(Forma.descricao) == db.func.lower(descricao)).first():
            flash(f'Erro: A forma "{descricao}" já está cadastrada.', 'danger')
        else:
            nova_forma = Forma(descricao=descricao)
            db.session.add(nova_forma)
            db.session.commit()
            flash('Forma adicionada com sucesso!', 'success')
        return redirect(url_for('gerenciar_formas'))
    
    # --- LÓGICA ATUALIZADA DE BUSCA E PAGINAÇÃO ---
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '', type=str)
    
    # A consulta agora só busca formas ativas
    query = Forma.query.filter_by(is_active=True)
    if q:
        query = query.filter(Forma.descricao.ilike(f'%{q}%'))

    paginacao = query.order_by(Forma.descricao).paginate(page=page, per_page=5, error_out=False)
    formas_da_pagina = paginacao.items
    
    return render_template('formas.html', formas=formas_da_pagina, paginacao=paginacao, q=q)


@app.route('/configuracoes', methods=['GET', 'POST'])
@login_required
def gerenciar_configuracoes():
    config = Configuracao.query.get(1)
    if not config:
        config = Configuracao(id=1)
        db.session.add(config)
        db.session.commit()
    if request.method == 'POST':
        config.salario_desejado = float(request.form.get('salario_desejado', 0) or 0)
        config.custo_gas = float(request.form.get('custo_gas', 0) or 0)
        config.custo_luz = float(request.form.get('custo_luz', 0) or 0)
        config.custo_agua = float(request.form.get('custo_agua', 0) or 0)
        config.custo_outros = float(request.form.get('custo_outros', 0) or 0)
        config.horas_por_dia = float(request.form.get('horas_por_dia', 0) or 0)
        config.dias_por_semana = int(request.form.get('dias_por_semana', 0) or 0)
        config.calcular_custo_hora()
        db.session.commit()
        flash('Configurações salvas e custo da hora atualizado!', 'success')
        return redirect(url_for('gerenciar_configuracoes'))
    return render_template('configuracoes.html', config=config)

# Em app.py

# (Lembre-se de ter a função normalize_text e os imports necessários acima)

@app.route('/fichas-tecnicas', methods=['GET', 'POST'])
@login_required
def gerenciar_fichas_tecnicas():
    if request.method == 'POST':
        nome_original = request.form.get('nome', '').strip()

        # --- Bloco de Validação Adicionado ---
        if not nome_original:
            flash('O nome da ficha técnica não pode estar em branco.', 'danger')
            return redirect(url_for('gerenciar_fichas_tecnicas'))

        nome_normalizado = normalize_text(nome_original)
        
        ficha_existente = FichaTecnica.query.filter(FichaTecnica.nome_normalizado == nome_normalizado).first()
        if ficha_existente:
            flash(f'Uma ficha técnica com o nome "{nome_original}" já existe.', 'danger')
            return redirect(url_for('gerenciar_fichas_tecnicas'))
        # --- Fim do Bloco de Validação ---

        # Cria a nova ficha se a validação passar
        nova_ficha = FichaTecnica(
            nome=nome_original,
            nome_normalizado=nome_normalizado, # Salva a versão normalizada
            rendimento=request.form.get('rendimento'),
            observacoes=request.form.get('observacoes')
        )
        db.session.add(nova_ficha)
        db.session.commit()
        
        flash('Ficha Técnica criada com sucesso!', 'success')
        # Redireciona para a página de detalhes da nova ficha, que é mais útil
        return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=nova_ficha.id))

    # --- Lógica para GET (exibição da lista) ---
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '', type=str)
    
    query = FichaTecnica.query
    
    # Busca agora é feita no campo normalizado para ignorar acentos
    if q:
        q_normalizado = f"%{normalize_text(q)}%"
        query = query.filter(FichaTecnica.nome_normalizado.ilike(q_normalizado))
        
    paginacao = query.order_by(FichaTecnica.nome).paginate(page=page, per_page=5, error_out=False)
    fichas_da_pagina = paginacao.items
    
    return render_template('fichas_tecnicas.html', 
                           fichas_tecnicas=fichas_da_pagina, 
                           paginacao=paginacao, 
                           q=q)

# Em app.py

@app.route('/ficha-tecnica/<int:ficha_tecnica_id>', methods=['GET', 'POST'])
@login_required
def detalhe_ficha_tecnica(ficha_tecnica_id):
    ficha = FichaTecnica.query.get_or_404(ficha_tecnica_id)
    todos_ingredientes = Ingrediente.query.order_by(Ingrediente.nome).all()
    
    # CORREÇÃO: Removida a linha duplicada que definia 'todas_formas'
    todas_formas = Forma.query.order_by(Forma.descricao).all()

    if request.method == 'POST':
        # Verifica se o formulário de "Salvar Detalhes" foi enviado
        if 'atualizar_ficha' in request.form:
            ficha.rendimento = request.form.get('rendimento')
            ficha.observacoes = request.form.get('observacoes')
            
            peso_str = request.form.get('peso_final_gramas')
            ficha.peso_final_gramas = float(peso_str) if peso_str else None

            ficha.tempo_producao_horas = int(request.form.get('tempo_producao_horas', 0) or 0)
            ficha.tempo_producao_minutos = int(request.form.get('tempo_producao_minutos', 0) or 0)
            
            ficha.incluir_custo_mao_de_obra = 'incluir_custo_mao_de_obra' in request.form
            
            ids_formas_selecionadas = request.form.getlist('formas_selecionadas')
            ficha.formas = Forma.query.filter(Forma.id.in_(ids_formas_selecionadas)).all()
            
            db.session.commit()
            flash('Ficha Técnica atualizada com sucesso!', 'success')
        
        # Verifica se o formulário de "Adicionar Ingrediente" foi enviado
        elif 'ingrediente_id' in request.form:
            nova_associacao = FichaTecnicaIngrediente(
                fichatecnica_id=ficha.id, 
                ingrediente_id=int(request.form['ingrediente_id']), 
                quantidade_usada=float(request.form['quantidade_usada'])
            )
            db.session.add(nova_associacao)
            db.session.commit()
            flash('Ingrediente adicionado à ficha!', 'success')

        return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=ficha.id))

    # Lógica para GET (exibição da página)
    page_ing = request.args.get('page_ing', 1, type=int)
    paginacao_ingredientes = FichaTecnicaIngrediente.query.filter_by(fichatecnica_id=ficha.id).paginate(page=page_ing, per_page=5, error_out=False)
    
    # CORREÇÃO: Apontando para o nome de arquivo correto que você usa
    return render_template(
        'ficha_tecnica_detalhe.html', 
        ficha=ficha, 
        todos_ingredientes=todos_ingredientes, 
        todas_formas=todas_formas, 
        paginacao_ingredientes=paginacao_ingredientes
    )



@app.route('/ficha-tecnica/produzir/<int:ficha_tecnica_id>', methods=['POST'])
@login_required
def produzir_ficha_tecnica(ficha_tecnica_id):
    ficha = FichaTecnica.query.get_or_404(ficha_tecnica_id)
    try:
        quantidade_a_produzir = int(request.form.get('quantidade_a_produzir', 1))
        if quantidade_a_produzir <= 0: raise ValueError()
    except (ValueError, TypeError):
        flash('A quantidade a produzir deve ser um número inteiro maior que zero.', 'warning')
        return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=ficha_tecnica_id))
    ingredientes_insuficientes = [item.ingrediente.nome for item in ficha.ingredientes if item.ingrediente.quantidade_estoque < (item.quantidade_usada * quantidade_a_produzir)]
    if ingredientes_insuficientes:
        flash(f'Não foi possível produzir. Estoque insuficiente para: {", ".join(ingredientes_insuficientes)}.', 'danger')
        return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=ficha_tecnica_id))
    for item in ficha.ingredientes:
        item.ingrediente.quantidade_estoque -= (item.quantidade_usada * quantidade_a_produzir)
    db.session.commit()
    flash(f'{quantidade_a_produzir} unidade(s) de "{ficha.nome}" produzida(s) com sucesso! Estoque atualizado.', 'success')
    return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=ficha_tecnica_id))

@app.route('/ficha-tecnica/ingrediente/delete/<int:item_id>', methods=['POST'])
@login_required
def deletar_item_ficha(item_id):
    item = FichaTecnicaIngrediente.query.get_or_404(item_id)
    ficha_id = item.fichatecnica_id
    db.session.delete(item)
    db.session.commit()
    flash('Ingrediente removido da ficha.', 'success')
    return redirect(url_for('detalhe_ficha_tecnica', ficha_tecnica_id=ficha_id))

@app.route('/ficha-tecnica/delete/<int:ficha_tecnica_id>', methods=['POST'])
@login_required
def deletar_ficha_tecnica(ficha_tecnica_id):
    ficha = FichaTecnica.query.get_or_404(ficha_tecnica_id)
    db.session.delete(ficha)
    db.session.commit()
    flash(f'Ficha Técnica "{ficha.nome}" deletada com sucesso.', 'warning')
    return redirect(url_for('gerenciar_fichas_tecnicas'))

@app.route('/exportar/ficha-tecnica/<int:ficha_tecnica_id>')
@login_required
def exportar_ficha_tecnica(ficha_tecnica_id):
    ficha = FichaTecnica.query.get_or_404(ficha_tecnica_id)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Ficha Técnica'
    bold_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right')
    sheet.merge_cells('A1:E1')
    title_cell = sheet['A1']
    title_cell.value = f"Ficha Técnica - {ficha.nome}"
    title_cell.font = title_font
    title_cell.alignment = center_align
    sheet.row_dimensions[1].height = 25
    sheet['A3'] = "Rendimento:"
    sheet['A3'].font = bold_font
    sheet['B3'] = ficha.rendimento
    sheet['A4'] = "Formas Utilizadas:"
    sheet['A4'].font = bold_font
    sheet['B4'] = ", ".join([forma.descricao for forma in ficha.formas])
    sheet.append(['']); sheet.append([''])
    headers = ['Ingrediente', 'Marca', 'Quantidade Usada', 'Unidade', 'Custo do Item (R$)']
    sheet.append(headers)
    for cell in sheet[7]:
        cell.font = bold_font
    for item in ficha.ingredientes:
        custo_item = item.ingrediente.preco_por_unidade() * item.quantidade_usada
        sheet.append([item.ingrediente.nome, item.ingrediente.marca, item.quantidade_usada, item.ingrediente.unidade_medida, round(custo_item, 2)])
    sheet.append([''])
    summary_start_row = sheet.max_row + 1
    if ficha.peso_final_gramas:
        sheet[f'A{summary_start_row}'] = "Rendimento Estimado:"
        sheet[f'A{summary_start_row}'].font = bold_font
        fatias = ficha.calcular_porcoes(100)
        sheet[f'B{summary_start_row}'] = f'{fatias:.1f} fatias de 100g'.replace('.',',')
    sheet[f'D{summary_start_row}'] = "CUSTO TOTAL:"
    sheet[f'D{summary_start_row}'].font = bold_font
    sheet[f'E{summary_start_row}'] = round(ficha.custo_total(), 2)
    sheet[f'E{summary_start_row}'].font = bold_font
    sheet[f'E{summary_start_row}'].alignment = right_align
    sheet.append(['']); sheet.append([''])
    obs_start_row = sheet.max_row
    sheet[f'A{obs_start_row}'] = "Observações / Modo de Preparo:"
    sheet[f'A{obs_start_row}'].font = bold_font
    sheet.merge_cells(f'A{obs_start_row + 1}:E{obs_start_row + 3}')
    obs_cell = sheet[f'A{obs_start_row + 1}']
    obs_cell.value = ficha.observacoes
    obs_cell.alignment = Alignment(wrap_text=True, vertical='top')
    sheet.column_dimensions['A'].width = 30; sheet.column_dimensions['B'].width = 20; sheet.column_dimensions['C'].width = 18; sheet.column_dimensions['D'].width = 15; sheet.column_dimensions['E'].width = 18
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"Ficha_Tecnica_{ficha.nome.replace(' ', '_')}.xlsx"
    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment;filename={filename}'})

# Em app.py

@app.route('/vendas', methods=['GET', 'POST'])
@login_required
def gerenciar_vendas():
    if request.method == 'POST':
        fichatecnica_id = request.form.get('fichatecnica_id')
        
        # Validação robusta para campos vazios e valores inválidos
        preco_venda_str = request.form.get('preco_venda_final')
        if not fichatecnica_id or not preco_venda_str:
            flash('Por favor, selecione uma ficha e preencha o preço de venda.', 'warning')
            return redirect(url_for('gerenciar_vendas'))
        
        try:
            quantidade = int(request.form.get('quantidade', 1))
            preco_venda = float(preco_venda_str)
            if preco_venda <= 0 or quantidade <= 0:
                raise ValueError("Valores devem ser positivos.")
        except (ValueError, TypeError):
            flash('Por favor, insira valores numéricos válidos para quantidade e preço.', 'danger')
            return redirect(url_for('gerenciar_vendas'))
        
        ficha = FichaTecnica.query.get(fichatecnica_id)
        if not ficha:
            flash('Ficha Técnica não encontrada.', 'danger')
            return redirect(url_for('gerenciar_vendas'))

        # Cálculos da venda
        custo_total_da_venda = ficha.custo_total() * quantidade
        lucro = preco_venda - custo_total_da_venda
        
        nova_venda = Venda(
            fichatecnica_id=fichatecnica_id, 
            quantidade=quantidade, 
            preco_venda_final=preco_venda, 
            custo_producao_total=custo_total_da_venda, 
            lucro_calculado=lucro
        )
        db.session.add(nova_venda)
        db.session.commit()
        flash('Venda registrada com sucesso!', 'success')
        return redirect(url_for('gerenciar_vendas'))

    # Lógica para exibir a página (GET)
    fichas_tecnicas = FichaTecnica.query.order_by(FichaTecnica.nome).all()
    vendas = Venda.query.order_by(Venda.data_venda.desc()).all()
    
    # Prepara um dicionário com os preços sugeridos para o JavaScript
    precos_sugeridos = {
        ficha.id: round(ficha.custo_total() * 2, 2)  # Custo x2 = 100% de margem
        for ficha in fichas_tecnicas
    }
    precos_sugeridos_json = json.dumps(precos_sugeridos)

    # Calcula os totais para o dashboard
    faturamento_total = sum(v.preco_venda_final for v in vendas)
    custo_total = sum(v.custo_producao_total for v in vendas)
    lucro_total = faturamento_total - custo_total

    return render_template('vendas.html', 
                           fichas_tecnicas=fichas_tecnicas, 
                           vendas=vendas,
                           faturamento_total=faturamento_total,
                           custo_total=custo_total,
                           lucro_total=lucro_total,
                           precos_sugeridos_json=precos_sugeridos_json)

# Em app.py

@app.route('/criar-primeiro-usuario-secreto-12345')
def criar_primeiro_usuario():
    # Verifica se já existe algum usuário
    if User.query.first():
        return 'Um usuário já existe. Esta rota não fará nada.'
    
    # Crie aqui o usuário com um nome e senha padrão
    username = 'admin'
    password = 'mudar_esta_senha_depois'
    
    new_user = User(username=username)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    
    return f'Usuário "{username}" criado com sucesso! Por segurança, remova esta rota do seu app.py agora.'





#orçamento

@app.route('/orcamento/novo')
@login_required
def novo_orcamento():
    """Mostra a página do formulário para criar um novo orçamento."""
    return render_template('gerador_orcamento.html')






@app.route('/orcamento/gerar-pdf', methods=['POST'])
@login_required
def gerar_pdf_orcamento():
    """Pega os dados do formulário e gera um orçamento em PDF."""
    
    dados_orcamento = request.form.to_dict()
    itens_json = request.form.get('itens_orcamento')
    dados_orcamento['itens'] = json.loads(itens_json) if itens_json else []

    if not dados_orcamento.get('nome_cliente') or not dados_orcamento['itens']:
        flash('Nome do cliente e pelo menos um item são obrigatórios.', 'warning')
        return redirect(url_for('novo_orcamento'))

    # --- INÍCIO DA CORREÇÃO DE DATA ---
    # Formata a data do pedido
    data_pedido_str = dados_orcamento.get('data_pedido')
    if data_pedido_str:
        dados_orcamento['data_pedido'] = datetime.strptime(data_pedido_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    
    # Formata a data do evento/entrega
    data_evento_str = dados_orcamento.get('data_evento')
    if data_evento_str:
        dados_orcamento['data_evento'] = datetime.strptime(data_evento_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    # --- FIM DA CORREÇÃO ---

    # Calcula os totais
    subtotal = sum(float(item['quantidade']) * float(item['valor_unitario']) for item in dados_orcamento['itens'])
    taxa_entrega = float(dados_orcamento.get('taxa_entrega', 0) or 0)
    dados_orcamento['subtotal'] = subtotal
    dados_orcamento['valor_total'] = subtotal + taxa_entrega
    
    # Renderiza o template HTML com os dados já formatados
    html_string = render_template('orcamento_pdf.html', **dados_orcamento)
    
    # Gera o PDF (código sem alteração)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=result)
    
    if pdf.err:
        flash(f'Ocorreu um erro ao gerar o PDF: {pdf.err}', 'danger')
        return redirect(url_for('novo_orcamento'))

    result.seek(0)
    nome_cliente = dados_orcamento.get('nome_cliente', 'cliente').replace(' ', '_')
    filename = f'orcamento_{nome_cliente}.pdf'
    
    return send_file(
        result,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

#para deletar as formas e editar.

# Em app.py

@app.route('/forma/delete/<int:id>', methods=['POST'])
@login_required
def deletar_forma(id):
    forma = Forma.query.get_or_404(id)
    # --- NOVA LÓGICA DE "EXCLUSÃO SUAVE" ---
    forma.is_active = False
    db.session.commit()
    flash('Forma excluída com sucesso! Ela não aparecerá para novas fichas.', 'success')
    return redirect(url_for('gerenciar_formas'))

@app.route('/forma/edit/<int:id>', methods=['POST'])
@login_required
def editar_forma(id):
    forma = Forma.query.get_or_404(id)
    nova_descricao = request.form.get('descricao', '').strip()
    if not nova_descricao:
        flash('A descrição não pode ser vazia.', 'danger')
        return redirect(url_for('gerenciar_formas'))
    
    outra_forma = Forma.query.filter(Forma.id != id, db.func.lower(Forma.descricao) == db.func.lower(nova_descricao)).first()
    if outra_forma:
        flash('Já existe outra forma com esta descrição.', 'danger')
    else:
        forma.descricao = nova_descricao
        db.session.commit()
        flash('Forma atualizada com sucesso!', 'success')
    return redirect(url_for('gerenciar_formas'))

@app.route('/setup-inicial-com-usuarios-12345')
def setup_inicial():
    try:
        with app.app_context():
            db.create_all()  # Cria todas as tabelas

            # Lista de usuários a serem criados
            usuarios_para_criar = [
                {'username': 'sabrina', 'password': '994507111'},
                {'username': 'neuza', 'password': '280588'}
            ]

            mensagens = []
            for user_info in usuarios_para_criar:
                # Verifica se o usuário já existe antes de criar
                if not User.query.filter_by(username=user_info['username']).first():
                    novo_usuario = User(username=user_info['username'])
                    novo_usuario.set_password(user_info['password'])
                    db.session.add(novo_usuario)
                    mensagens.append(f"Usuário '{user_info['username']}' criado com sucesso.")
                else:
                    mensagens.append(f"Usuário '{user_info['username']}' já existe.")

            db.session.commit()

            return f"<h1>Setup Concluído!</h1><ul><li>{'</li><li>'.join(mensagens)}</li></ul>"

    except Exception as e:
        return f"<h1>Erro no Setup</h1><p>{str(e)}</p>"



if __name__ == '__main__':
    app.run(debug=True)